#__author:  Administrator
#date:  2020/3/8

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    #打开文件
    file = load_workbook(filename=path)
    #返回所有表格的名称
    #print(file.get_sheet_names())
    sheets = file.sheetnames
    #拿出一个表格
    sheet = file[sheets[0]]
    #最大行数
    #print(sheet.max_row)
    #最大列数
    #print(sheet.max_column)
    #表名
    #print(sheet.title)

    #读取一张表的数据
    for lineNum in range(1,sheet.max_row +1):
        lineList = []
        for columnNum in range(1,sheet.max_column +1):
            #拿数据
            value = sheet.cell(row=lineNum,column=columnNum).value
            #if value != None:
            lineList.append(value)
            print(lineList)

#不能处理xls文件
path = r'/Users/admin/Desktop/成本-第三批优化.xlsx'
readXlsxFile(path)