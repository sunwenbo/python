#__author:  Administrator
#date:  2020/3/8
from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    dic = {}
    file = load_workbook(filename=path)
    sheets = file.sheetnames
    for sheetName in sheets:
        sheet = file[sheetName]
        #一张表的所有数据
        sheetInfo = []
        for lineNum in range(1, sheet.max_row + 1):
            lineList = []
            for columnNum in range(1,sheet.max_column +1):
                value = sheet.cell(row=lineNum,column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)
        #将一张表的数据存到字典
        dic[sheetName] = sheetInfo
    return dic
path = r'E:\python笔记\3.自动化办公\4.读取excel文件\加密机_密钥接口分析.xlsx'
dic = readXlsxFile(path)
print(dic["Sheet1"])
print(len(dic))