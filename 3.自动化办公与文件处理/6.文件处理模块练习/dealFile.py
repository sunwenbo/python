#__author:  Administrator
#date:  2020/3/11

import csv
import win32com
import win32com.client
from openpyxl.reader.excel import load_workbook
from collections import OrderedDict
from pyexcel_xls import get_data
from pyexcel_xls import save_data

class DealFile(object):
    def readCsv(self,path):
        with open(path,"r",encoding='utf-8') as f:
            infoList = []
            allFileInfo = csv.reader(f)
            for row in allFileInfo:
                infoList.append(row)
        return infoList

    def writeCsv(self,path, data):
        #writeCsv(path,[[1,2,3],[4,5,6],[7,8,9]])
        with open(path, "w", encoding="utf-8", newline="") as f:
            write = csv.writer(f)
            for rowData in data:
                write.writerow(rowData)

    def readWordFile(self,path):
        #调用系统word功能，可以处理doc和docx两种文件
        mw = win32com.client.Dispatch("Word.application")
        #打开文件
        doc = mw.Documents.Open(path)
        for par in doc.Paragraphs:
            lin = par.Range.Text
            print(lin)
        #关闭文件
        doc.Close()
        #退出word
        mw.Quit()
    def readWordFiletoOtherFile(self,path,topath):
        mw = win32com.client.Dispatch("Word.application")
        doc = mw.Documents.Open(path)
        #将word文件的内容保存到另一个文件
        doc.SaveAs(topath,2)  #2代表text文件
        doc.Close()
        mw.Quit()
    def makeWordFile(self,path):
        word = win32com.client.Dispatch("Word.Application")
        word.Visible = True
        doc = word.Documents.Add()
        r = doc.Range(0,0)
        r.InsertAfter("        插入文字\n")
        doc.SaveAs(path)
        doc.Close()
        word.Quit()

    def readXlsxFile(self,path):
        dic = {}
        file = load_workbook(filename=path)
        sheets = file.sheetnames
        for sheetName in sheets:
            sheet = file[sheetName]
            sheetInfo = []
            for lineNum in range(1, sheet.max_row + 1):
                lineList = []
                for columnNum in range(1, sheet.max_column + 1):
                    value = sheet.cell(row=lineNum, column=columnNum).value
                    lineList.append(value)
                sheetInfo.append(lineList)
            dic[sheetName] = sheetInfo
        return dic

    def readXlsAndXlsxFile(self,path):
        dic = OrderedDict()
        xdata = get_data(path)
        for sheet in xdata:
            dic[sheet] = xdata[sheet]
            return dic

    def makeExcelFile(self,path, data):
        dic = OrderedDict()
        for sheetName, sheetValue in data.items():
            d = {}
            d[sheetName] = sheetValue
            dic.update(d)
        save_data(path, dic)